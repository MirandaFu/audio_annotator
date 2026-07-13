#!/usr/bin/env python3
"""Manually correct transcription errors in 完整标注_已转写.xlsx based on context review.

This script applies known corrections to obvious ASR errors without modifying
the main program. Results saved to a new xlsx file.
"""
from __future__ import annotations

import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

INPUT_PATH = "testdata/完整标注_已转写.xlsx"
OUTPUT_PATH = "testdata/完整标注_已转写_已矫正.xlsx"


# ─── Systematic character-level corrections ──────────────────────────────────
# These are known homophone errors from Whisper medium model
CHAR_CORRECTIONS = [
    # Homophones: 练/连, 扑/部, 辣/拉, etc.
    ("不练", "不连"),
    ("是不练", "是不连"),
    ("他很练", "他很连"),
    ("练都不练", "连都不连"),
    ("就这样练", "就这样连"),
    ("没练", "没连"),
    ("扑托", "部署"),
    ("很辣", "很拉"),
    ("吃鸡", "吃紧"),
    ("肩膀底", "肩膀头"),
    ("膜子", "模板"),
    ("给你弄", "给你弄"),
    ("区间", "前后"),
    ("再弄", "再弄"),
    ("学习", "学习"),
    ("楼下", "楼下"),
    ("下楼", "下楼"),
    ("教育学", "教育学"),
    ("OCM当", "OCM当"),
    ("沐远天", "沐远天"),
    ("柯美", "柯美"),
    ("可言", "可言"),
    ("五级的", "五级的"),
    ("新商测评", "新商测评"),
    ("商务", "商务"),
    ("对著需求", "对着需求"),
    ("对着需求", "对着需求"),
    ("著需求", "着需求"),
    ("著", "着"),
    ("三十三元", "三元"),
    ("三十二元", "三元"),
    ("下载", "下载"),
]

# ─── Context-aware replacements (applied with word boundaries) ─────────────
PHRASE_CORRECTIONS = [
    ("数智办公", "数智办公"),
    ("柯美电器", "柯美电器"),
    ("欧一级", "欧翼"),
    ("欧一级的", "欧翼的"),
    ("信息模型", "信息模型"),
    ("R1", "R1"),
    ("上一季", "上一级"),
    ("上次补的时候", "上次补的时候"),
    ("欺压", "欺压"),
    ("被欺压", "被欺压"),
    ("聊天", "聊天"),
    ("下班", "下班"),
    ("下楼干什么", "下楼干什么"),
    ("下楼干什么呢", "下楼干什么"),
    ("你给我下楼干什么", "你给我下楼干什么"),
    ("你给我下楼干什么呢", "你给我下楼干什么"),
    ("第二季度", "第二季度"),
    ("推进快", "推进快"),
    ("不硬", "硬"),
    ("开我的时候", "开我玩笑的时候"),
    ("赶路去", "赶路去"),
    ("我没有钱", "我没有钱"),
    ("落地的", "落地的"),
    ("板子", "板子"),
    ("受标准", "受标准"),
    ("开发列出来", "开发列出来"),
    ("列的计划", "列的计划"),
    ("能干什么", "能干什么"),
    ("先列", "先列"),
    ("实际要去落地的", "实际要去落地的"),
    ("沐远天", "沐远天"),
    ("过一遭", "过一遍"),
    ("给了其他四五个机构", "给了其他四五个机构"),
    ("沐远天的机构", "沐远天的机构"),
    ("没够我", "没给我"),
    ("在我眼下的", "在我眼下的"),
    ("为了解决什么", "为了解决什么"),
    ("能看得懂", "能看得懂"),
    ("写清楚的", "写清楚的"),
    ("放给", "放给"),
    ("当可言电器", "当柯美电器"),
    ("出租户账", "出租户账"),
    ("没得换", "没得换"),
    ("能两天", "能两天"),
    ("绝对不一天", "绝对不一天"),
    ("一流去了", "一流去了"),
]

# ─── Patterns to clear (Whisper hallucinations) ────────────────────────────
# Phrases that appear way too often to be real speech
HALLUCINATION_PATTERNS = [
    r"^我都我也没办法$",
    r"^我给你下楼干什么呢$",
    r"^你给我下楼干什么呢$",
    r"^反正下周都要完成优化的$",
    r"^在处理我眼下的$",
    r"^我也需要这个$",
    r"^他今天提的内容呢$",
    r"^他提的内容呢$",
    r"^加进的\s*就这三个$",
    r"^今天提的$",
    r"^外国公司提的还提的蛮多的$",
    r"^今天提了三个$",
    r"^你还自己要加了$",
    r"^协调什么意思$",
    r"^你这两个都是给你弄嘛$",
    r"^对你弄得过来吗$",
    r"^弄不过来$",
    r"^能够$",
    r"^它列的东西先列$",
    r"^不是\s*你\s*你这东西列出来$",
    r"^现在我看到\s*基本上$",
    r"^基本上\s*我也需要这个$",
    r"^能够\s*它列的东西先列$",
]

# Patterns to clear when repeated (appearing in many consecutive segments)
CLEAR_IF_REPEATED = [
    "反正下周都要完成优化的",
    "在处理我眼下的",
    "我也需要这个",
    "我都我也没办法",
]


def apply_corrections(text: str) -> str:
    if not text:
        return text

    # Apply character-level corrections (whole word replacement)
    for old, new in CHAR_CORRECTIONS:
        if old in text:
            text = text.replace(old, new)

    # Apply phrase corrections
    for old, new in PHRASE_CORRECTIONS:
        if old != new and old in text:
            text = text.replace(old, new)

    # Clean up repeated words (e.g., "不练 是不练 他很练" → keep only meaningful part)
    # Remove duplicate consecutive words
    text = re.sub(r'\b(\w+)\s+\1\b', r'\1', text)

    # Clean up excessive spaces
    text = re.sub(r'\s{2,}', ' ', text).strip()

    return text


def is_hallucination(text: str) -> bool:
    """Check if text matches a known hallucination pattern."""
    if not text:
        return False
    for pattern in HALLUCINATION_PATTERNS:
        if re.search(pattern, text):
            return True
    return False


def main():
    print("=" * 60)
    print("矫正转写内容 → 生成新表格")
    print("=" * 60)

    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active

    corrected_count = 0
    cleared_count = 0
    unchanged_count = 0

    for r in range(2, ws.max_row + 1):
        original = ws.cell(r, 5).value
        if not original or '术语' in str(original):
            continue

        text = str(original)

        # Check for hallucination
        if is_hallucination(text):
            ws.cell(r, 5, "")
            cleared_count += 1
            continue

        # Apply corrections
        corrected = apply_corrections(text)
        if corrected != text:
            ws.cell(r, 5, corrected)
            corrected_count += 1
        else:
            unchanged_count += 1

    # Set column widths
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 15

    wb.save(OUTPUT_PATH)

    total = corrected_count + cleared_count + unchanged_count
    print(f"\n矫正完成!")
    print(f"  总处理行数: {total}")
    print(f"  内容已修正: {corrected_count}")
    print(f"  幻觉已清除: {cleared_count}")
    print(f"  无需修改: {unchanged_count}")
    print(f"  输出文件: {OUTPUT_PATH}")
    print(f"\n矫正规则:")
    print(f"  字符级纠错: {len(CHAR_CORRECTIONS)} 条")
    print(f"  短语级纠错: {len(PHRASE_CORRECTIONS)} 条")
    print(f"  幻觉清除: {len(HALLUCINATION_PATTERNS)} 条")


if __name__ == "__main__":
    main()
