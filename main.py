"""
Audio Annotator — 会议音频说话人标注工具
用法: python main.py [audio.wav]
"""
import sys
import os
import time
import threading
import numpy as np
import sounddevice as sd
import soundfile as sf
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from waveform_widget import WaveformWidget
from segments_table import SegmentsTable
from speaker_panel import SpeakerPanel


SPEAKER_COLORS = [
    "#E74C3C", "#3498DB", "#2ECC71", "#F39C12",
    "#9B59B6", "#1ABC9C", "#E67E22", "#34495E",
    "#E84393", "#00B894", "#6C5CE7", "#FDCB6E",
]


def get_duration(wav_path):
    import subprocess
    result = subprocess.run(
        ["ffprobe", "-v", "error",
         "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1",
         wav_path],
        capture_output=True, text=True
    )
    try:
        return float(result.stdout.strip())
    except ValueError:
        return None


def load_wav_data(wav_path):
    data, fs = sf.read(wav_path, dtype="float32", always_2d=False)
    if len(data.shape) > 1:
        data = data.mean(axis=1)
    return data.astype(np.float32), fs


CHUNK = 2048


class AudioEngine:
    """Audio playback via sounddevice with wall-clock position tracking."""

    def __init__(self):
        self._data = None
        self._fs = 48000
        self._channels = 1
        self._stream = None
        self._thread = None
        self._stop_evt = threading.Event()
        self._pos_callback = None
        self._root = None
        self._start_wall = 0.0
        self._start_pos = 0.0
        self._duration = 0.0
        self._timer = None
        self._eof = threading.Event()
        self._stream_ready = threading.Event()
        self._play_error = None
        self._error_callback = None
        self._output_device = None  # None = use sounddevice default
        self._play_offset = 0

    def load(self, path):
        self.stop()
        data, fs = sf.read(path, dtype="float32", always_2d=False)
        if len(data.shape) > 1:
            data = data.mean(axis=1)
        self._data = data.astype(np.float32)
        self._fs = fs
        self._duration = len(data) / fs
        self._channels = 1 if len(data.shape) == 1 else data.shape[1]

    @property
    def duration(self):
        return self._duration

    @property
    def sample_rate(self):
        return self._fs

    @staticmethod
    def get_output_devices():
        devices = []
        try:
            for i, d in enumerate(sd.query_devices()):
                if d["max_output_channels"] > 0:
                    devices.append((i, d["name"]))
        except Exception:
            pass
        return devices

    @staticmethod
    def find_best_output_device():
        """Prefer a real speaker device over virtual audio devices."""
        VIRTUAL_HINTS = ("virtual", "oray", "blackhole", "soundflower", "loopback",
                         "audio hijack", "voicemeeter", "cable", "apowermirror")
        devices = AudioEngine.get_output_devices()
        # First pass: non-virtual devices
        for idx, name in devices:
            if not any(h in name.lower() for h in VIRTUAL_HINTS):
                return idx, name
        # Fallback: first available device
        if devices:
            return devices[0]
        return None, None

    @property
    def current_time(self):
        if self._stream is not None and self._stream.active:
            return min(self._start_pos + (time.perf_counter() - self._start_wall), self._duration)
        return self._start_pos

    def is_playing(self):
        return self._stream is not None and self._stream.active

    def set_position_callback(self, cb, root=None):
        self._root = root
        self._pos_callback = cb

    def set_error_callback(self, cb):
        self._error_callback = cb

    def _report_error(self, msg):
        self._play_error = msg
        if self._error_callback:
            try:
                self._root.after_idle(self._error_callback, msg)
            except Exception:
                pass

    def _notify(self):
        if self._pos_callback:
            try:
                self._root.after_idle(self._pos_callback, self.current_time)
            except Exception:
                pass

    def _tick(self):
        if self._eof.is_set() or self.current_time >= self._duration:
            self._eof.clear()
            if self._pos_callback:
                try:
                    self._pos_callback(self.current_time)
                except Exception:
                    pass
            self._stop_timer()
            return
        if self._stream is None or not self._stream.active:
            self._stop_timer()
            return
        if self._pos_callback:
            try:
                self._pos_callback(self.current_time)
            except Exception:
                pass
        self._timer = self._root.after(4, self._tick)

    def _stop_timer(self):
        if self._timer:
            try:
                self._root.after_cancel(self._timer)
            except Exception:
                pass
            self._timer = None

    def _stream_callback(self, outdata, frames, time_info, status):
        if status:
            pass  # Ignore underflow/overflow warnings during normal operation
        if self._stop_evt.is_set():
            outdata[:] = 0
            return sd.CallbackStop
        offset = self._play_offset
        end = offset + frames
        available = len(self._data) - offset
        if available <= 0:
            outdata[:] = 0
            self._eof.set()
            return sd.CallbackStop
        n = min(frames, available)
        block = self._data[offset:offset + n]
        if self._channels == 1:
            outdata[:n, 0] = block
            if n < frames:
                outdata[n:, 0] = 0
        else:
            outdata[:n] = block
            if n < frames:
                outdata[n:] = 0
        self._play_offset = offset + n
        if self._play_offset >= len(self._data):
            self._eof.set()
            return sd.CallbackStop
        return None  # Continue streaming

    def _play_loop(self):
        try:
            if self._output_device is None:
                dev = None  # Follow system default
            else:
                dev = self._output_device
            self._play_offset = int(self._start_pos * self._fs)
            stream = sd.OutputStream(
                samplerate=self._fs,
                device=dev,
                channels=self._channels,
                dtype="float32",
                blocksize=128,
                latency='low',
                callback=self._stream_callback,
            )
            stream.start()
            self._stream = stream
            self._stream_ready.set()
            while stream.active and not self._stop_evt.is_set():
                self._stop_evt.wait(0.05)
        except Exception as e:
            self._report_error(str(e))
        finally:
            self._eof.set()
            self._stream_ready.set()
            s = self._stream
            self._stream = None
            if s:
                try:
                    s.abort()
                except Exception:
                    pass
                try:
                    s.close()
                except Exception:
                    pass

    def _cleanup(self):
        self._stop_evt.set()
        self._stream_ready.set()  # Unblock play() if it's waiting
        self._stop_timer()
        s = self._stream
        self._stream = None
        if s:
            try:
                s.abort()
            except Exception:
                pass
            try:
                s.close()
            except Exception:
                pass
        if self._thread:
            self._thread.join(timeout=1.0)
            self._thread = None

    def play(self, pos=None):
        self._cleanup()
        if self._data is None or len(self._data) == 0:
            return
        if pos is not None:
            self._start_pos = max(0.0, min(pos, self._duration))
        self._start_wall = time.perf_counter()
        self._stop_evt.clear()
        self._eof.clear()
        self._stream_ready.clear()
        self._play_error = None
        self._thread = threading.Thread(target=self._play_loop, daemon=True)
        self._thread.start()
        if self._stream_ready.wait(timeout=2.0):
            if self._play_error:
                self._cleanup()
                return
            self._tick()
        else:
            self._cleanup()

    def pause(self):
        if not self.is_playing():
            return
        self._start_pos = self.current_time
        self._cleanup()

    def resume(self):
        if self.is_playing():
            return
        self.play()

    def stop(self):
        self._cleanup()
        self._start_pos = 0.0

    def seek(self, t):
        was_playing = self.is_playing()
        self._cleanup()
        self._start_pos = max(0.0, min(t, self._duration))
        if was_playing:
            self._start_wall = time.perf_counter()
            self._stop_evt.clear()
            self._eof.clear()
            self._stream_ready.clear()
            self._play_error = None
            self._thread = threading.Thread(target=self._play_loop, daemon=True)
            self._thread.start()
            if self._stream_ready.wait(timeout=2.0):
                if self._play_error:
                    self._cleanup()
                    return
                self._tick()
            else:
                self._cleanup()


class AudioAnnotator:
    def __init__(self, root, audio_path=None):
        self.root = root
        self.root.title("Audio Annotator — 说话人标注工具")
        self.root.geometry("1400x800")

        self.audio_path = audio_path
        self.audio_data = None
        self.sample_rate = 48000
        self.duration = 0.0
        self.current_time = 0.0
        self.playing = False
        self.audio = AudioEngine()
        self.audio.set_error_callback(self._on_audio_error)
        self.segments = []
        self.current_speaker = None
        self.drag_start = None
        self._seeking = False

        self._build_ui()
        if audio_path and os.path.exists(audio_path):
            self.load_audio(audio_path)

    def _build_ui(self):
        # Row 1: Transport + device + time + speaker
        row1 = ttk.Frame(self.root)
        row1.pack(fill="x", padx=6, pady=(6, 2))

        ttk.Button(row1, text="📂 打开", command=self.open_file).pack(side="left", padx=2)
        self.btn_play = ttk.Button(row1, text="▶ 播放", command=self._on_play_click, width=8)
        self.btn_play.pack(side="left", padx=2)
        ttk.Button(row1, text="⏹ 从头播放", command=self.stop, width=6).pack(side="left", padx=2)

        ttk.Separator(row1, orient="vertical").pack(side="left", fill="y", padx=6)
        self.lbl_time = ttk.Label(row1, text="00:00:00.0 / 00:00:00.0", font=("Courier", 10))
        self.lbl_time.pack(side="left", padx=4)

        ttk.Separator(row1, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Label(row1, text="输出:").pack(side="left")
        self.device_var = tk.StringVar()
        self._device_map = {}
        self.combo_device = ttk.Combobox(row1, textvariable=self.device_var,
                                         state="readonly", width=18)
        self.combo_device.pack(side="left", padx=2)
        self.combo_device.bind("<<ComboboxSelected>>", self._on_device_change)
        self._refresh_device_list()

        ttk.Separator(row1, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Label(row1, text="发言人:").pack(side="left")
        self.lbl_current = ttk.Label(row1, text="说话人1",
                                     foreground="#E74C3C", font=("Helvetica", 10, "bold"))
        self.lbl_current.pack(side="left", padx=4)

        # Row 2: Annotation tools + zoom
        row2 = ttk.Frame(self.root)
        row2.pack(fill="x", padx=6, pady=(2, 4))

        self.btn_mark = tk.Checkbutton(row2, text="📌 打点模式", indicatoron=False,
                                       selectcolor="#4a7c59", command=self._toggle_mark_mode)
        self.btn_mark.pack(side="left", padx=2)
        ttk.Button(row2, text="💾 导出", command=self.export_segments).pack(side="left", padx=2)
        ttk.Button(row2, text="🗑 删除", command=self._delete_selected).pack(side="left", padx=2)

        ttk.Separator(row2, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Label(row2, text="缩放:").pack(side="left")
        ttk.Button(row2, text="＋", command=self.zoom_in, width=3).pack(side="left", padx=1)
        self.lbl_zoom = ttk.Label(row2, text="1.0x", width=5)
        self.lbl_zoom.pack(side="left")
        ttk.Button(row2, text="－", command=self.zoom_out, width=3).pack(side="left", padx=1)
        ttk.Button(row2, text="⟲", command=self.zoom_reset, width=3).pack(side="left", padx=1)

        ttk.Separator(row2, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Label(row2, text="高度:").pack(side="left")
        self.scale_height = ttk.Scale(row2, from_=0.5, to=3.0, value=1.0,
                                       length=70, command=self._on_height_change)
        self.scale_height.pack(side="left", padx=2)

        seek_frame = ttk.Frame(self.root)
        seek_frame.pack(fill="x", padx=6, pady=(0, 4))
        self.seek_var = tk.DoubleVar(value=0)
        self.seek_scale = ttk.Scale(seek_frame, from_=0, to=100, variable=self.seek_var,
                                     command=self._on_seek_drag)
        self.seek_scale.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self.seek_scale.bind("<ButtonPress-1>", lambda e: setattr(self, '_seeking', True))
        self.seek_scale.bind("<ButtonRelease-1>", self._on_seek_release)
        self.lbl_seek = ttk.Label(seek_frame, text="00:00:00.0 / 00:00:00.0", width=22)
        self.lbl_seek.pack(side="right")

        main_paned = ttk.PanedWindow(self.root, orient="horizontal")
        main_paned.pack(fill="both", expand=True, padx=6, pady=4)

        left_paned = ttk.PanedWindow(main_paned, orient="vertical")
        main_paned.add(left_paned, weight=4)

        waveform_container = ttk.Frame(left_paned)
        self.waveform = WaveformWidget(
            waveform_container,
            on_play_at=self._seek_to,
            on_drag_start=self._on_drag_start,
            on_drag_end=self._on_drag_end,
            on_mark_complete=self._on_mark_complete,
        )
        self._waveform_scrollbar = tk.Scrollbar(waveform_container, orient="horizontal")
        self._waveform_scrollbar.config(command=self.waveform._on_scrollbar)
        self.waveform.scrollbar = self._waveform_scrollbar

        self.waveform.bind_to(self.root)
        self.root.bind("<MouseWheel>", self._on_root_mousewheel)
        self.root.bind("<Button-4>", self._on_root_mousewheel)
        self.root.bind("<Button-5>", self._on_root_mousewheel)

        self.waveform.pack(fill="both", expand=True)
        left_paned.add(waveform_container, weight=3)

        table_container = ttk.Frame(left_paned)
        left_paned.add(table_container, weight=1)

        ttk.Label(table_container, text="标注片段 (双击编辑说话人 | Delete 删除 | 双击波形播放)",
                  font=("Helvetica", 9)).pack(anchor="w", pady=2)
        self.table = SegmentsTable(
            table_container,
            on_select_time=self._seek_to,
            on_delete=self._delete_segment,
            on_edit=self._refresh_table,
        )
        self.table.pack(fill="both", expand=True)

        speaker_container = ttk.Frame(main_paned)
        main_paned.add(speaker_container, weight=1)

        self.current_speaker = "说话人1"

        self.speaker_panel = SpeakerPanel(
            speaker_container,
            speakers=["说话人1", "说话人2"],
            colors={"说话人1": SPEAKER_COLORS[0], "说话人2": SPEAKER_COLORS[1]},
            current=self.current_speaker,
            on_select=self._on_speaker_select,
            on_rename=self._on_speaker_rename,
            on_add=self._on_speaker_add,
            on_delete=self._on_speaker_delete,
        )
        self.speaker_panel.pack(fill="both", expand=True)

        self.status = ttk.Label(self.root, text="就绪 — 请打开 WAV 文件开始标注", relief="sunken", anchor="w")
        self.status.pack(fill="x", side="bottom")

    # ─── Audio device ──────────────────────────────────────────────────────

    def _refresh_device_list(self):
        devices = AudioEngine.get_output_devices()
        self._device_map = {}
        values = ["系统默认"]  # None = follow system default
        self._device_map["系统默认"] = None
        for idx, name in devices:
            self._device_map[name] = idx
            values.append(name)
        self.combo_device["values"] = values
        self.device_var.set("系统默认")
        self._on_device_change()

    def _on_device_change(self, _=None):
        name = self.device_var.get()
        idx = self._device_map.get(name)
        self.audio._output_device = idx

    # ─── Audio I/O ─────────────────────────────────────────────────────────

    def open_file(self):
        path = filedialog.askopenfilename(
            parent=self.root,
            title="选择音频文件",
            filetypes=[("WAV 音频", "*.wav"), ("所有文件", "*.*")],
        )
        if path:
            self.load_audio(path)

    def load_audio(self, path):
        self.stop()
        self.audio_path = path
        self.status.config(text=f"正在加载: {os.path.basename(path)} ...")
        self.root.update_idletasks()
        try:
            self.audio.load(path)
            self.audio_data = self.audio._data
            self.sample_rate = self.audio.sample_rate
            self.duration = self.audio.duration
        except Exception as e:
            messagebox.showerror("加载失败", f"无法读取音频文件:\n{e}")
            self.status.config(text="加载失败")
            return
        self.waveform.set_audio(self.audio_data, self.sample_rate)
        self.current_time = 0.0
        self._update_time_label()
        if self.duration > 0:
            self.seek_scale.config(to=100)
        self.status.config(text=f"已加载: {os.path.basename(path)}  |  "
                                f"时长: {self._fmt(self.duration)}  |  "
                                f"采样率: {self.sample_rate} Hz")

    def play(self):
        if not self.audio_path or self.audio._data is None:
            return
        if self.playing:
            self.pause()
            return
        self.playing = True
        self.audio.set_position_callback(self._on_pos_update, self.root)
        self.audio.play(self.current_time)
        self._update_play_button()

    def pause(self):
        if not self.playing:
            return
        self.audio.pause()
        self.playing = False
        self.current_time = self.audio.current_time
        self.waveform.set_playhead(self.current_time)
        self._update_play_button()
        self._update_time_label()

    def resume(self):
        if self.playing:
            return
        self.playing = True
        self.audio.resume()
        self._update_play_button()

    def stop(self):
        self.playing = False
        self.audio.stop()
        self.current_time = 0.0
        self.waveform.set_playhead(0.0)
        self._update_time_label()
        self._update_play_button()

    def _on_pos_update(self, t):
        if t >= self.duration and self.playing:
            self.playing = False
            self.current_time = 0.0
            self.waveform.set_playhead(0.0)
            self._update_play_button()
            self._update_time_label()
            return
        self.current_time = t
        self.waveform.set_playhead(t)
        self._update_time_label()

    def _on_audio_error(self, msg):
        self.playing = False
        self._update_play_button()
        self.status.config(text=f"播放错误: {msg}")
        try:
            messagebox.showerror("音频播放错误", f"无法播放音频:\n{msg}", parent=self.root)
        except Exception:
            pass

    def _on_play_click(self):
        if self.playing:
            self.pause()
        else:
            self.play()

    # ─── Segments ──────────────────────────────────────────────────────────

    def _on_drag_start(self, x, t):
        self.drag_start = (x, t)

    def _on_drag_end(self, x, t):
        if self.drag_start is None:
            return
        _, t0 = self.drag_start
        self.drag_start = None
        t_start = min(t0, t)
        t_end = max(t0, t)
        if t_end - t_start < 0.05:
            return
        speaker = self.current_speaker or "说话人1"
        seg = {"start": t_start, "end": t_end, "speaker": speaker}
        self.segments.append(seg)
        self.segments.sort(key=lambda s: s["start"])
        self._refresh_table()

    def _on_speaker_select(self, name):
        self.current_speaker = name
        color = self.speaker_panel.get_colors().get(name, "#000")
        self.lbl_current.config(text=name, foreground=color)
        self.waveform.set_current_speaker(name)
        self.speaker_panel.set_current(name)

    def _on_speaker_rename(self, old_name, new_name):
        if not self.speaker_panel.rename_speaker(old_name, new_name):
            return
        for seg in self.segments:
            if seg["speaker"] == old_name:
                seg["speaker"] = new_name
        self._refresh_table()
        if self.current_speaker == old_name:
            self.current_speaker = new_name
            color = self.speaker_panel.get_colors().get(new_name, "#000")
            self.lbl_current.config(text=new_name, foreground=color)
            self.waveform.set_current_speaker(new_name)
            self.speaker_panel.set_current(new_name)

    def _on_speaker_add(self, name):
        self._on_speaker_select(name)
        self.status.config(text=f"已添加发言人: {name}")

    def _on_speaker_delete(self, name):
        new_current = self.speaker_panel.current
        for seg in self.segments:
            if seg["speaker"] == name:
                seg["speaker"] = new_current
        self._refresh_table()
        self.current_speaker = new_current
        color = self.speaker_panel.get_colors().get(new_current, "#000")
        self.lbl_current.config(text=new_current, foreground=color)
        self.waveform.set_current_speaker(new_current)
        self.speaker_panel.set_current(new_current)
        self.status.config(text=f"已删除发言人: {name}")

    def zoom_in(self):
        self.waveform.zoom_in()
        self._update_zoom_label()

    def zoom_out(self):
        self.waveform.zoom_out()
        self._update_zoom_label()

    def zoom_reset(self):
        self.waveform.zoom_reset()
        self._update_zoom_label()

    def _update_zoom_label(self):
        self.lbl_zoom.config(text=f"{self.waveform._zoom:.1f}x")

    def _on_root_mousewheel(self, ev):
        if hasattr(self, 'waveform') and self.waveform:
            self.waveform._on_mousewheel(ev)

    def _on_height_change(self, val):
        self.waveform.set_height_scale(float(val))

    def _on_seek_drag(self, val):
        if self.duration <= 0:
            return
        t = float(val) / 100.0 * self.duration
        self.current_time = max(0, min(t, self.duration))
        self.waveform.set_playhead(self.current_time)
        self._update_time_label()

    def _on_seek_release(self, _):
        self._seeking = False
        if self.duration > 0:
            self._seek_to(self.current_time)

    def _delete_selected(self):
        sel = self.table.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在下方列表中选择一个片段")
            return
        item = sel[0]
        idx = self.table.tree.index(item)
        if 0 <= idx < len(self.segments):
            del self.segments[idx]
            self._refresh_table()
            self.status.config(text=f"已删除片段 #{idx+1}")

    def _seek_to(self, t):
        was_playing = self.playing
        self.current_time = max(0, min(t, self.duration))
        self.waveform.set_playhead(self.current_time)
        self._update_time_label()
        self.audio.seek(self.current_time)
        if was_playing:
            self.playing = True
            self._update_play_button()

    def _toggle_mark_mode(self):
        active = self.btn_mark.cget('text')
        if active == '📌 打点模式':
            self.waveform.set_mark_mode(True)
            self.btn_mark.config(text='📌 打点中...')
            self.status.config(text="打点模式：单击波形设置起点，再单击设置终点，自动弹出发言人选择")
        else:
            self.waveform.set_mark_mode(False)
            self.btn_mark.config(text='📌 打点模式')
            self.status.config(text="已退出打点模式")

    def _on_mark_complete(self, start, end):
        popup = tk.Toplevel(self.root)
        popup.title("选择发言人")
        popup.geometry("300x160")
        popup.transient(self.root)
        popup.grab_set()

        ttk.Label(popup, text=f"片段: {self._fmt(start)} → {self._fmt(end)}").pack(pady=8)

        ttk.Label(popup, text="选择发言人:").pack(anchor="w", padx=12)
        speakers = self.speaker_panel.get_speakers()
        var = tk.StringVar(value=self.current_speaker or speakers[0] if speakers else "")
        combo = ttk.Combobox(popup, textvariable=var, values=speakers,
                             state="normal", width=20)
        combo.pack(padx=12, pady=4)
        combo.focus_set()

        def confirm():
            speaker = var.get().strip()
            if not speaker:
                speaker = speakers[0] if speakers else "说话人1"
            seg = {"start": start, "end": end, "speaker": speaker}
            self.segments.append(seg)
            self.segments.sort(key=lambda s: s["start"])
            self._refresh_table()
            self.status.config(text=f"已标注: {speaker} [{self._fmt(start)} - {self._fmt(end)}]")
            popup.destroy()
            self._toggle_mark_mode()

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=8)
        ttk.Button(btn_frame, text="确定", command=confirm).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="取消", command=lambda: (popup.destroy(), self._toggle_mark_mode())).pack(side="left", padx=4)
        combo.bind("<Return>", lambda e: confirm())

    def _delete_segment(self, index):
        if 0 <= index < len(self.segments):
            del self.segments[index]
            self._refresh_table()

    def _refresh_table(self):
        colors = self.speaker_panel.get_colors()
        self.table.set_segments(self.segments, colors)
        self.waveform.set_segments(self.segments, colors)

    def export_segments(self):
        if not self.segments:
            messagebox.showinfo("提示", "没有标注片段可导出")
            return
        base = os.path.splitext(os.path.basename(self.audio_path or "标注"))[0]

        path = filedialog.asksaveasfilename(
            parent=self.root,
            title="导出标注文件",
            defaultextension=".txt",
            initialfile=base + "_标注.txt",
            filetypes=[
                ("文本文件 (Tab分隔)", "*.txt"),
                ("CSV 表格 (可导入Excel)", "*.csv"),
                ("Excel 表格 (.xlsx)", "*.xlsx"),
                ("所有文件", "*.*"),
            ],
        )
        if not path:
            return

        ext = os.path.splitext(path)[1].lower()

        if ext == ".xlsx":
            self._export_xlsx(path)
            return

        is_csv = ext == ".csv"

        lines = []
        if is_csv:
            lines.append("讲话人,开始时间,结束时间,时长")
            sep = ","
        else:
            sep = "\t"

        for seg in self.segments:
            s = self._fmt(seg["start"])
            e = self._fmt(seg["end"])
            dur = self._fmt(max(0, seg["end"] - seg["start"]))
            speaker = seg["speaker"]
            if is_csv and ("," in speaker or '"' in speaker):
                speaker = '"' + speaker.replace('"', '""') + '"'
            lines.append(f"{speaker}{sep}{s}{sep}{e}{sep}{dur}")

        try:
            encoding = "utf-8-sig" if is_csv else "utf-8"
            with open(path, "w", encoding=encoding) as f:
                f.write("\n".join(lines) + "\n")
            self.status.config(text=f"已导出 {len(self.segments)} 个片段 → {path}")
            messagebox.showinfo("导出成功", f"已导出 {len(self.segments)} 个标注片段\n格式: {'CSV表格' if is_csv else '文本'}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _export_xlsx(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "标注结果"

        headers = ["讲话人", "开始时间", "结束时间", "时长"]
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for row_idx, seg in enumerate(self.segments, 2):
            speaker = seg["speaker"]
            s = self._fmt(seg["start"])
            e = self._fmt(seg["end"])
            dur = self._fmt(max(0, seg["end"] - seg["start"]))

            ws.cell(row=row_idx, column=1, value=speaker)
            ws.cell(row=row_idx, column=2, value=s)
            ws.cell(row=row_idx, column=3, value=e)
            ws.cell(row=row_idx, column=4, value=dur)

            for col in range(1, 5):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 14

        try:
            wb.save(path)
            self.status.config(text=f"已导出 {len(self.segments)} 个片段 → {path}")
            messagebox.showinfo("导出成功", f"已导出 {len(self.segments)} 个标注片段\n格式: Excel表格")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _update_play_button(self):
        if self.playing:
            self.btn_play.config(text="⏸ 暂停")
        else:
            self.btn_play.config(text="▶ 播放")

    def _update_time_label(self):
        if getattr(self, '_updating_label', False):
            return
        self._updating_label = True
        try:
            cur = self._fmt(self.current_time)
            dur = self._fmt(self.duration)
            self.lbl_time.config(text=f"{cur} / {dur}")
            self.lbl_seek.config(text=f"{cur} / {dur}")
            if self.duration > 0 and not self._seeking:
                self.seek_var.set(self.current_time / self.duration * 100)
        finally:
            self._updating_label = False

    @staticmethod
    def _fmt(t):
        if t is None or t < 0:
            t = 0
        h = int(t // 3600)
        m = int((t % 3600) // 60)
        s = t % 60
        return f"{h:02d}:{m:02d}:{s:05.2f}"


def _macos_open_file(title="选择文件", file_types=None):
    if file_types is None:
        file_types = ["wav"]
    type_list = ", ".join(f'"{t}"' for t in file_types)
    script = f'POSIX path of (choose file of type {{{type_list}}} with prompt "{title}")'
    try:
        import subprocess
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass
    return None


def _macos_save_file(default_name="export.txt", title="保存文件"):
    script = f'POSIX path of (choose file name with prompt "{title}" default name "{default_name}")'
    try:
        import subprocess
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass
    return None


def _parse_time(t_str):
    parts = t_str.strip().split(":")
    if len(parts) != 3:
        raise ValueError(f"Invalid time format: {t_str}")
    h = int(parts[0])
    m = int(parts[1])
    s = float(parts[2])
    return h * 3600 + m * 60 + s


def _cli_export(audio_path, export_path):
    base = os.path.splitext(audio_path)[0]
    annot_path = base + "_标注.txt"

    if not os.path.exists(annot_path):
        print(f"Error: annotation file not found: {annot_path}")
        print("Create a file named <audio_basename>_标注.txt with format:")
        print("  讲话人<tab>开始时间<tab>结束时间")
        sys.exit(1)

    segments = []
    with open(annot_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) >= 3:
                segments.append({
                    "speaker": parts[0],
                    "start": parts[1],
                    "end": parts[2],
                })

    if not segments:
        print("Error: no segments found in annotation file")
        sys.exit(1)

    ext = os.path.splitext(export_path)[1].lower()
    is_csv = ext == ".csv"

    lines = []
    if is_csv:
        lines.append("讲话人,开始时间,结束时间,时长")
        sep = ","
    else:
        sep = "\t"

    for seg in segments:
        speaker = seg["speaker"]
        if is_csv and ("," in speaker or '"' in speaker):
            speaker = '"' + speaker.replace('"', '""') + '"'
        start_s = seg["start"]
        end_s = seg["end"]
        try:
            start_t = _parse_time(start_s)
            end_t = _parse_time(end_s)
            dur = max(0, end_t - start_t)
            dur_str = AudioAnnotator._fmt(dur)
        except (ValueError, TypeError):
            dur_str = end_s  # fallback
        lines.append(f"{speaker}{sep}{start_s}{sep}{end_s}{sep}{dur_str}")

    encoding = "utf-8-sig" if is_csv else "utf-8"
    with open(export_path, "w", encoding=encoding) as f:
        f.write("\n".join(lines) + "\n")

    print(f"Exported {len(segments)} segments to: {export_path}")
    for seg in segments:
        start_t = _parse_time(seg["start"])
        end_t = _parse_time(seg["end"])
        dur = max(0, end_t - start_t)
        print(f"  {seg['speaker']}\t{seg['start']}\t{seg['end']}\t{AudioAnnotator._fmt(dur)}")


def main():
    audio_path = sys.argv[1] if len(sys.argv) > 1 else None

    if "--export" in sys.argv:
        idx = sys.argv.index("--export")
        export_path = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else None
        if audio_path and export_path:
            _cli_export(audio_path, export_path)
        return

    root = tk.Tk()
    app = AudioAnnotator(root, audio_path)
    root.mainloop()


if __name__ == "__main__":
    main()
