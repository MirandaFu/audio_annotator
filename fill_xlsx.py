#!/usr/bin/env python3
"""Transcribe audio and fill text into 完整标注.xlsx time segments."""
from __future__ import annotations

import sys
import os
import time
import json

sys.stdout.reconfigure(line_buffering=True)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from correction import CorrectionConfig, load_config
from transcriber import (
    FasterWhisperTranscriber,
    transcribe_audio_segment,
    match_transcription_to_segments,
    VadConfig,
    merge_short_segments,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "testdata", "完整标注.xlsx")
AUDIO_PATH = os.path.join(BASE_DIR, "testdata", "3-数智办公四月第二次例会.wav")
OUTPUT_PATH = os.path.join(BASE_DIR, "testdata", "完整标注_已转写.xlsx")
CONFIG_PATH = os.path.join(BASE_DIR, "testdata", "correction-config.json")


def time_str_to_float(ts: str) -> float:
    """Convert HH:MM:SS.ff to seconds."""
    parts = ts.strip().split(":")
    if len(parts) == 3:
        h, m, rest = parts
        s, *ms = rest.split(".")
        sec = float(s) + (float("0." + ms[0]) if ms else 0.0)
        return int(h) * 3600 + int(m) * 60 + sec
    elif len(parts) == 2:
        m, rest = parts
        s, *ms = rest.split(".")
        sec = float(s) + (float("0." + ms[0]) if ms else 0.0)
        return int(m) * 60 + sec
    return float(ts)


def load_xlsx_segments(path: str) -> list[dict]:
    """Load annotation segments from xlsx."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    segments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        speaker, start_str, end_str, duration = row
        if start_str is None:
            continue
        segments.append({
            "speaker": str(speaker or ""),
            "start": time_str_to_float(str(start_str)),
            "end": time_str_to_float(str(end_str)),
            "duration": str(duration or ""),
        })
    return segments


def main():
    t0 = time.time()
    print("=" * 60)
    print("音频转写 → 填入标注表格")
    print("=" * 60)

    # 1. Load segments
    print(f"\n[1/5] 读取标注文件: {XLSX_PATH}")
    segments = load_xlsx_segments(XLSX_PATH)
    print(f"  共 {len(segments)} 个标注片段")
    if segments:
        print(f"  时间范围: {segments[0]['start']:.2f}s → {segments[-1]['end']:.2f}s")

    # 2. Load correction config
    print(f"\n[2/5] 加载纠错配置: {CONFIG_PATH}")
    config = load_config(CONFIG_PATH)
    print(f"  用户词典: {len(config.custom_terms)} 个术语")
    print(f"  正则规则: {len(config.regex_rules)} 条")

    # 3. Transcription
    print(f"\n[3/5] 转写音频 (模型: medium, float16)...")
    print(f"  音频文件: {AUDIO_PATH}")
    if not os.path.exists(AUDIO_PATH):
        print(f"  ERROR: 音频文件不存在: {AUDIO_PATH}")
        sys.exit(1)

    vad_cfg = VadConfig(
        min_silence_duration_ms=800,
        speech_pad_ms=300,
        preset="conservative",
    )

    print("  加载模型 (首次需下载 ~2.5GB)...")
    transcriber = FasterWhisperTranscriber(model_size="medium", device="cpu", compute_type="default")

    prompt = config.build_initial_prompt()
    generic = "这是一段中文会议录音，包含专业术语和项目名称。"
    if prompt.startswith(generic[:20]) and len(config.custom_terms) == 0:
        prompt = None

    print("  全文件转写中...")
    transcribed, info, word_data = transcriber.transcribe_with_timestamps(
        AUDIO_PATH,
        language="zh",
        initial_prompt=prompt,
        word_timestamps=True,
        vad_config=vad_cfg,
    )
    print(f"  转写完成: {len(transcribed)} 个文本段")

    # 4. Skip merging - use raw VAD segments for better matching granularity
    print("\n[4/5] 使用原始转写片段 (不合并)...")
    print(f"  转写片段: {len(transcribed)} 个文本段")

    # 5. Match to annotation segments
    print("\n[5/5] 匹配标注片段...")
    matches = match_transcription_to_segments(transcribed, segments)

    updated = 0
    empty = 0
    for idx, (seg, text) in enumerate(matches):
        corrected = config.apply(text)
        seg["text"] = corrected
        if corrected:
            updated += 1
        else:
            empty += 1
        if (idx + 1) % 100 == 0 or idx == len(matches) - 1:
            print(f"  进度: {idx + 1}/{len(matches)} (有内容: {updated}, 无内容: {empty})")

    # 6. Write to new xlsx
    print(f"\n写入输出文件: {OUTPUT_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Add "内容" header
    ws.cell(1, 5, "内容")
    ws.cell(1, 5).font = Font(bold=True, size=11)
    ws.cell(1, 5).alignment = Alignment(horizontal="center", vertical="center")

    # Add word confidence info column
    ws.cell(1, 6, "置信度")
    ws.cell(1, 6).font = Font(bold=True, size=11)
    ws.cell(1, 6).alignment = Alignment(horizontal="center", vertical="center")

    for i, seg in enumerate(segments):
        row = i + 2
        ws.cell(row, 5, seg.get("text", ""))
        ws.cell(row, 5).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Auto-width for column E
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 15

    wb.save(OUTPUT_PATH)

    elapsed = time.time() - t0
    print(f"\n{'=' * 60}")
    print(f"完成! 用时: {elapsed:.1f}s")
    print(f"总片段: {len(segments)}")
    print(f"有转写内容: {updated}")
    print(f"无转写内容: {empty}")
    print(f"输出文件: {OUTPUT_PATH}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
